import sys
import os
import io
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import tkinter.font as tkFont
import subprocess
import MatchTable
import openpyxl

title = '対局くん'

root = tk.Tk()
root.title(title)
root.resizable(False, False)

fontStyle = tkFont.Font(family='游ゴシック', size=12)

# 入力エリアフレーム作成
input_frame = ttk.Frame(root, padding=(32))
input_frame.grid()

# エクセルファイル名指定のEntry
label1 = ttk.Label(input_frame, text='エクセルファイル名', padding=(5, 2), font=fontStyle)
label1.grid(row=0, column=0, sticky=tk.W)

excel_file_name = tk.StringVar()
excel_file_name_entry = ttk.Entry(
    input_frame,
    textvariable=excel_file_name,
    width=72,
    font=fontStyle)
excel_file_name_entry.grid(row=0, column=1, sticky=tk.W, padx=4)

excel_file_name_button = tk.Button(
    input_frame,
    text='参照',
    command=lambda: excel_file_name_select_proc(),
    font=fontStyle,
    width=6)
excel_file_name_button.grid(row=0, column=2, sticky=tk.W, padx=4)

# シート名指定のEntry
label2 = ttk.Label(input_frame, text='シート名', padding=(5, 2), font=fontStyle)
label2.grid(row=1, column=0, sticky=tk.W)

sheetname_select = tk.StringVar()
sheetname_combobox = ttk.Combobox(
    input_frame,
    textvariable=sheetname_select,
    values=[],
    width=30,
    state='readonly',
    font=fontStyle)
sheetname_combobox.grid(row=1, column=1, sticky=tk.W, padx=4)

# コマンドのCombobox
label3 = ttk.Label(input_frame, text='コマンド', padding=(5, 2), font=fontStyle)
label3.grid(row=2, column=0, sticky=tk.W)

cmds = ['組み合わせ作成(登録順で組み合わせ）', '組み合わせ作成(スコア・棋力差・初期ソート順で組み合わせ）', '成績作成']
cmd_select = tk.StringVar()
cmd_combobox = ttk.Combobox(
    input_frame,
    textvariable=cmd_select,
    values=cmds,
    width=70,
    state='readonly',
    font=fontStyle)
cmd_combobox.set(cmds[0])
cmd_combobox.bind(
    '<<ComboboxSelected>>',
    lambda e: cmd_select_proc())
cmd_combobox.grid(row=2, column=1, sticky=tk.W, padx=4)

# 対局回数のEntry
label4 = ttk.Label(input_frame, text='回戦数', padding=(5, 2), font=fontStyle)
label4.grid(row=3, column=0, sticky=tk.W)

taikyoku_kaisu = tk.IntVar()
taikyoku_kaisu_entry = ttk.Entry(
    input_frame,
    textvariable=taikyoku_kaisu,
    width=4,
    font=fontStyle)
taikyoku_kaisu_entry.grid(row=3, column=1, sticky=tk.W, padx=4)

# ボタン配置フレーム作成
button_frame = ttk.Frame(input_frame, padding=(0, 5))
button_frame.grid(row=4, column=1, sticky=tk.W, padx=4)

excel_run_button = tk.Button(
    button_frame,
    text='エクセルで見る',
    command=lambda: execute_excel_proc(),
    font=fontStyle,
    width=14)
excel_run_button.pack(side=tk.LEFT, padx=4)

match_table_button = tk.Button(
    button_frame,
    text='コマンド処理を行う',
    command=lambda: match_table_proc(),
    font=fontStyle,
    width=14)
match_table_button.pack(side=tk.LEFT, padx=4)

quit_button = tk.Button(
    button_frame,
    text='終了',
    command=lambda: quit_proc(),
    font=fontStyle,
    width=14)
quit_button.pack(side=tk.LEFT, padx=4)

# エクセルファイル参照ボタン選択時処理
def excel_file_name_select_proc():
    fTyp = [("","*.xlsx")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    file = filedialog.askopenfilename(filetypes = fTyp, initialdir = iDir)
    excel_file_name.set(file)

    # シート名設定
    wb = openpyxl.load_workbook(excel_file_name.get())
    sheetname_combobox.config(values=wb.sheetnames)
    sheetname_combobox.set(wb.sheetnames[0])

# コマンド選択時処理
def cmd_select_proc():
    if cmd_select.get() == '成績作成':
        taikyoku_kaisu_entry.config(state="disabled")
        taikyoku_kaisu.set(0)
    else:
        taikyoku_kaisu_entry.config(state="normal")

# エクセル起動処理
def execute_excel_proc():
    if excel_file_name.get() == '':
        messagebox.showinfo(title, "エクセルファイル名を入力してください。")
        return
    subprocess.Popen(['start', excel_file_name.get()], shell=True)

# MatchTable起動処理
def match_table_proc():
    if excel_file_name.get() == '':
        messagebox.showinfo(title, "エクセルファイル名を入力してください。")
        return
    if sheetname_select.get() == '':
        messagebox.showinfo(title, "シート名を入力してください。")
        return

    proc = MatchTable.MatchTable()

    if cmd_select.get() == '成績作成':
        # 成績作成処理の結果、エラーがあれば標準出力に内容が出るのでそれをリダイレクトして取得。
        with io.StringIO() as s:
            sys.stdout = s
            proc.write_result(excel_file_name.get(), sheetname_select.get(), excel_file_name.get())
            contents = s.getvalue()
        if contents == '':
            messagebox.showinfo(title, "成績の作成が完了しました。")
        else:
            messagebox.showinfo(title, "成績の作成でエラーが発生しました。エラー内容は以下の通りです。\n\n{}".format(contents))
        execute_excel_proc()

    elif cmd_select.get() == '組み合わせ作成(登録順で組み合わせ）':

        # 組み合わせ作成
        taisenNo = taikyoku_kaisu.get()
        if taisenNo < 1:
            messagebox.showinfo(title, "回戦数を入力してください。")
            return
        # 組み合わせ処理の結果、エラーがあれば標準出力に内容が出るのでそれをリダイレクトして取得。
        with io.StringIO() as s:
            sys.stdout = s
            proc.player_decision2(taisenNo, excel_file_name.get(), sheetname_select.get(), excel_file_name.get())
            contents = s.getvalue()
        if contents == '':
            messagebox.showinfo(title, "組み合わせの作成が完了しました。")
        else:
            messagebox.showinfo(title, "組み合わせの作成でエラーが発生しました。エラー内容は以下の通りです。\n\n{}".format(contents))
        execute_excel_proc()

    else:

        # 組み合わせ作成
        taisenNo = taikyoku_kaisu.get()
        if taisenNo < 1:
            messagebox.showinfo(title, "回戦数を入力してください。")
            return
        # 組み合わせ処理の結果、エラーがあれば標準出力に内容が出るのでそれをリダイレクトして取得。
        with io.StringIO() as s:
            sys.stdout = s
            proc.player_decision(taisenNo, excel_file_name.get(), sheetname_select.get(), excel_file_name.get())
            contents = s.getvalue()
        if contents == '':
            messagebox.showinfo(title, "組み合わせの作成が完了しました。")
        else:
            messagebox.showinfo(title, "組み合わせの作成でエラーが発生しました。エラー内容は以下の通りです。\n\n{}".format(contents))
        execute_excel_proc()

# 終了処理
def quit_proc():
    root.destroy()

root.mainloop()