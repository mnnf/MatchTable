import sys
import openpyxl
import datetime
import random
from collections import namedtuple
import zenhan
import math

# 参加者のデータ構造体定義
class taisensha_info:
    def __init__(self, row, random_row, pair_no, name, kiryoku, kiryoku_random_val, seibetsu, score, taisen_rireki, no_battles_cnt):
        self.row = row
        self.pair_no = pair_no
        self.kiryoku_random_val = kiryoku_random_val
        self.name = name
        self.kiryoku = kiryoku
        self.seibetsu = seibetsu
        self.score = score
        self.taisen_rireki = taisen_rireki
        self.sos = 0
        self.sosos = 0
        self.jyuni = 0
        self.no_battles_cnt = no_battles_cnt
        self.random_row = random_row

# 参加者の対戦履歴のデータ構造体定義
class taisen_rireki:
    def __init__(self, no, name, pair_name, taisensha_name1, taisensha_name2, kekka):
        self.no = no
        self.name = name
        self.pair_name = pair_name
        self.taisensha_name1 = taisensha_name1
        self.taisensha_name2 = taisensha_name2
        self.kekka = kekka

# 対戦相手のデータ構造体定義
class taisen_aite_info:
    def __init__(self, taisensha_info, taikyoku_su, index):
        self.taisensha_info = taisensha_info
        self.taikyoku_su = taikyoku_su
        self.index = index

# ペア碁マッチングクラス
class PairMatchTable:
    def __init__(self, seibetsu_flag, pair_kotei_flag, kiryoku_random_val_size):

        # 性別でペアを組むならTrue
        self.seibetsu_flag = seibetsu_flag

        # ペアを固定で組むならTrue
        self.pair_kotei_flag = pair_kotei_flag

        # 乱数での最終順序
        self.random_row_size = 100

        # 乱数での棋力変動幅
        self.kiryoku_random_val_size = kiryoku_random_val_size

        # タイトルの行位置
        self.title_row = 1

        # 参加者リストの開始行位置
        self.start_sankasha_row = 3

        # 参加者リストの名前列位置
        self.sankasha_name_col = 1

        # 対戦回の開始列位置
        self.taisen_start_col = 4

        # 勝ち数の列位置
        self.WIN_col = 0

        # SOSの列位置
        self.SOS_col = 0

        # SOSOSの列位置
        self.SOSOS_col = 0

        # 順位の列位置
        self.JYUNI_col = 0

        # 参加者リスト
        self.taisensha_info_list = []

        # 相手先ペアリスト
        self.aite_pair_info_list = []

        # 対戦番号(１～）
        self.taisenNo = 0

    # 棋力を取得(5D=5,1D=1,1K=0,2K=-1)
    def get_kiryoku(self, dat):
        kiryoku = 0
        if dat[1:2] == 'D':
            kiryoku = int(dat.replace('D', ''))
        else:
            kiryoku = 1 - int(dat.replace('K', ''))
        return kiryoku

    # 対戦未決定リスト取得
    def get_taisen_mikettei_list(self, taisensha_info_rec, pair_info_rec, step):

        mikettei_list = []

        for rec in self.taisensha_info_list:
            if rec.name == taisensha_info_rec.name:
                continue
            if rec.name == pair_info_rec.name:
                continue
            # 既に対戦番号(1～)以上の対局を実施した人は対象外
            if len(rec.taisen_rireki) >= self.taisenNo:
                continue
            # 過去に対戦した人は対象外
            if step == 1:
                past_flag = False
                for rireki in taisensha_info_rec.taisen_rireki:
                    if rireki.taisensha_name1 == rec.name:
                        past_flag = True
                        break
                    if rireki.taisensha_name2 == rec.name:
                        past_flag = True
                        break
                if past_flag:
                    continue
            # 相手は同性でなければ対象外
            if self.seibetsu_flag:
                if taisensha_info_rec.seibetsu != rec.seibetsu:
                    continue
            # それ以外を候補とする
            mikettei_list.append(rec)

        return mikettei_list

    # ペア未決定リスト取得
    def get_pair_mikettei_list(self, taisensha_info_rec, step):

        mikettei_list = []

        for rec in self.aite_pair_info_list:
            # 自分自身は対象外
            if rec.name == taisensha_info_rec.name:
                continue
            # 既に対戦番号(1～)以上の対局を実施した人は対象外
            if len(rec.taisen_rireki) >= self.taisenNo:
                continue
            # 過去にペアを組んだ人は対象外
            if step == 1:
                past_pairs_flag = False
                for rireki in taisensha_info_rec.taisen_rireki:
                    if rireki.pair_name == rec.name:
                        past_pairs_flag = True
                        break
                if past_pairs_flag:
                    continue
            # 同性は対象外
            if self.seibetsu_flag:
                if taisensha_info_rec.seibetsu == rec.seibetsu:
                    continue
            # それ以外を候補とする
            mikettei_list.append(rec)

        return mikettei_list

    # 対戦相手のペア未決定リスト取得
    def get_aite_pair_mikettei_list(self, taisensha_info_rec, pair_info_rec, taisen_aite_info_rec, step):

        mikettei_list = []

        for rec in self.aite_pair_info_list:
            # 自分自身は対象外
            if rec.name == taisensha_info_rec.name:
                continue
            # 相手ペアは既に選択されているので対象外
            if rec.name == pair_info_rec.name:
                continue
            # 対戦相手は対象外
            if rec.name == taisen_aite_info_rec.name:
                continue
            # 既に対戦番号(1～)以上の対局を実施した人は対象外
            if len(rec.taisen_rireki) >= self.taisenNo:
                continue
            # 過去にペアを組んだ人は対象外
            if step == 1:
                past_pairs_flag = False
                for rireki in taisen_aite_info_rec.taisen_rireki:
                    if rireki.pair_name == rec.name:
                        past_pairs_flag = True
                        break
                if past_pairs_flag:
                    continue
            # 同性は対象外
            if self.seibetsu_flag:
                if taisensha_info_rec.seibetsu == rec.seibetsu:
                    continue
            # それ以外を候補とする
            mikettei_list.append(rec)

        return mikettei_list

    # 対戦者を決定
    def get_taisen_kettei_sub(self, taisensha_info_rec, pair_info_rec, step):

        # 対戦者未決定リスト取得
        mikettei_list = self.get_taisen_mikettei_list(taisensha_info_rec, pair_info_rec, step)

        # 対戦者なし
        if len(mikettei_list) == 0:
            return None

        mikettei_list2 = []

        for index, aite_info in enumerate(mikettei_list):
            # 過去対戦数を取得
            taikyoku_su = self.get_taikyoku_su(taisensha_info_rec.taisen_rireki, aite_info.name)
            mikettei_list2.append(taisen_aite_info(aite_info, taikyoku_su, index))

        # 対戦回数・登録順にする
        mikettei_list2 = sorted(mikettei_list2, key=lambda x: (x.taikyoku_su, x.index))

        return mikettei_list2[0].taisensha_info

    # 対戦者を決定
    def get_taisen_kettei(self, taisensha_info_rec, pair_info_rec):

        aite_info = self.get_taisen_kettei_sub(taisensha_info_rec, pair_info_rec, step = 1)

        if aite_info != None:
            return aite_info

        aite_info = self.get_taisen_kettei_sub(taisensha_info_rec, pair_info_rec, step = 2)

        return aite_info

    # 固定ペアを取得
    def get_kotei_pair(self, taisensha_info_rec):
        for rec in self.taisensha_info_list:
            if rec.name == taisensha_info_rec.name:
                continue
            if rec.pair_no == taisensha_info_rec.pair_no:
                return rec
        return None

    # ペアを決定
    def get_pair_kettei_sub(self, taisensha_info_rec, step):

        if self.pair_kotei_flag:
            return get_kotei_pair(taisensha_info_rec)

        # ペア未決定リスト取得
        mikettei_list = self.get_pair_mikettei_list(taisensha_info_rec, step)

        # 対戦者なし
        if len(mikettei_list) == 0:
            return None

        return mikettei_list[0]

    # ペアを決定
    def get_pair_kettei(self, taisensha_info_rec):

        pair_rec = self.get_pair_kettei_sub(taisensha_info_rec, step = 1)
        if pair_rec != None:
            return pair_rec

        pair_rec = self.get_pair_kettei_sub(taisensha_info_rec, step = 2)
        return pair_rec

    # 対戦相手のペアを決定
    def get_aite_pair_kettei_sub(self, taisensha_info_rec, pair_info_rec, taisen_aite_info_rec, step):

        if self.pair_kotei_flag:
            return self.get_kotei_pair(taisen_aite_info_rec)        

        # ペア未決定リスト取得
        mikettei_list = self.get_aite_pair_mikettei_list(taisensha_info_rec, pair_info_rec, taisen_aite_info_rec, step)

        # 対戦者なし
        if len(mikettei_list) == 0:
            return None

        return mikettei_list[0]

    # 対戦相手のペアを決定
    def get_aite_pair_kettei(self, taisensha_info_rec, pair_info_rec, taisen_aite_info_rec):

        aite_pair_rec = self.get_aite_pair_kettei_sub(taisensha_info_rec, pair_info_rec, taisen_aite_info_rec, step = 1)
        if aite_pair_rec != None:
            return aite_pair_rec

        aite_pair_rec = self.get_aite_pair_kettei_sub(taisensha_info_rec, pair_info_rec, taisen_aite_info_rec, step = 2)

        return aite_pair_rec

    # スコアを取得
    def get_score(self, taisen_rireki_list):

        score = 0
        for rec in taisen_rireki_list:
            if (rec.kekka == '〇'):
                score += 1
        return score

    # 過去の対局数
    def get_taikyoku_su(self, taisen_rireki_list, taisen_aite_name):
        taikyoku_su = 0
        for rec in taisen_rireki_list:
            if rec.taisensha_name1 == taisen_aite_name:
                taikyoku_su += 1
            if rec.taisensha_name2 == taisen_aite_name:
                taikyoku_su += 1
        return taikyoku_su

    # 対戦相手の対戦者情報を取得
    def get_aite_info(self, name):
        for rec in self.taisensha_info_list:
            if rec.name == name:
                return rec
        return None

    # SOSを計算
    def get_sos(self, taisensha_info_rec):
        sos = 0
        for rec in taisensha_info_rec.taisen_rireki:
            aite_info = self.get_aite_info(rec.taisensha_name1)
            if aite_info != None:
                score = self.get_score(aite_info.taisen_rireki)
                sos += score
            aite_info = self.get_aite_info(rec.taisensha_name2)
            if aite_info != None:
                score = self.get_score(aite_info.taisen_rireki)
                sos += score
        return sos

    # SOSOSを計算
    def get_sosos(self, taisensha_info_rec):
        sosos = 0
        for rec in taisensha_info_rec.taisen_rireki:
            aite_info = self.get_aite_info(rec.taisensha_name1)
            if aite_info != None:
                sos = self.get_sos(aite_info)
                sosos += sos
            aite_info = self.get_aite_info(rec.taisensha_name2)
            if aite_info != None:
                sos = self.get_sos(aite_info)
                sosos += sos
        return sosos

    # 成績の列位置を取得
    def result_col_info(self, sheet):
        for col in range(1, sheet.max_column + 1):
            title = sheet.cell(self.title_row, col).value
            if title == '勝ち数':
                self.WIN_col = col
            if title == 'SOS':
                self.SOS_col = col
            if title == 'SOSOS':
                self.SOSOS_col = col
            if title == '順位':
                self.JYUNI_col = col

    # エクセルから参加者と過去の対戦情報を読み取り
    def read_excel(self, sheet):

        for row in range(self.start_sankasha_row, sheet.max_row + 1):

            # 対局者名を取得
            name = sheet.cell(row, self.sankasha_name_col).value
            if name != None:

                # ランダム数を取得
                random_row = random.randint(1, self.random_row_size)

                # ペア固定時に使用するペア番号を取得
                pair_no = math.floor((row - self.start_sankasha_row) / 2 + 1)

                # 性別を取得
                seibetsu = sheet.cell(row, self.sankasha_name_col + 1).value

                # 棋力を取得
                kiryoku = self.get_kiryoku(sheet.cell(row, self.sankasha_name_col + 2).value)

                # 棋力ランダム数を取得
                kiryoku_random_val = random.randint(-self.kiryoku_random_val_size, self.kiryoku_random_val_size)

                no_battles_cnt = 0

                # 対戦履歴情報取得
                taisen_rireki_info_list = []

                for taisenNo_loop in range(1, self.taisenNo + 1):
                    # ペアの名前、対戦者の名前と結果を取得
                    col = self.taisen_start_col + (taisenNo_loop - 1) * 4
                    pair_name = sheet.cell(row, col).value
                    kekka = sheet.cell(row, col + 3).value
                    if kekka != None:
                        # 対戦者履歴に追加
                        if pair_name == None:
                            # 不戦勝
                            no_battles_cnt += 1
                            taisen_rireki_info = taisen_rireki(no = taisenNo_loop, name = name, pair_name = None, taisensha_name1 = None, taisensha_name2 = None, kekka = kekka)
                            taisen_rireki_info_list.append(taisen_rireki_info)
                        else:
                            taisensha_names = sheet.cell(row, col + 1).value.split('、')
                            taisensha_name1 = taisensha_names[0]
                            taisensha_name2 = taisensha_names[1]
                            taisen_rireki_info = taisen_rireki(no = taisenNo_loop, name = name, pair_name = pair_name, taisensha_name1 = taisensha_name1, taisensha_name2 = taisensha_name2, kekka = kekka)
                            taisen_rireki_info_list.append(taisen_rireki_info)

                # 参加者の勝ち星を取得
                score = self.get_score(taisen_rireki_info_list)

                # 参加者リストに追加
                self.taisensha_info_list.append(taisensha_info(row = row, random_row = random_row, pair_no = pair_no, name = name, kiryoku = kiryoku, kiryoku_random_val = kiryoku_random_val, seibetsu = seibetsu, score = score, taisen_rireki = taisen_rireki_info_list, no_battles_cnt = no_battles_cnt))

    # 過去の対戦情報の矛盾をチェック
    def check_taisen_rireki(self, last_taisen_no):
        error_flag = False
        for rec in self.taisensha_info_list:
            for i, senreki in enumerate(rec.taisen_rireki):
                if i < last_taisen_no:
                    aite_info = self.get_aite_info(senreki.taisensha_name1)
                    if aite_info == None:
                        # 不戦勝などは対局者情報が取得できない。
                        continue
                    if len(aite_info.taisen_rireki) > i:
                        aite_senreki = aite_info.taisen_rireki[i]
                        if senreki.kekka == aite_senreki.kekka:
                            print('{}回戦の {} vs {} 戦の結果が両者同じです。'.format(i+1, rec.name, aite_info.name))
                            error_flag = True
        return not error_flag

    # ハンディキャップ取得
    def get_handycap(self, name1, name2, name3, name4):

        kiryoku1 = (self.get_aite_info(name1).kiryoku + 
                    self.get_aite_info(name2).kiryoku) / 2

        kiryoku2 = (self.get_aite_info(name3).kiryoku + 
                    self.get_aite_info(name4).kiryoku) / 2

        handy_diff = kiryoku1 - kiryoku2

        msg = ''
        if handy_diff == 0:
            msg = '互先 コミ６目'
        else:
            point = (handy_diff / 0.5)
            if point > 0:
                if point >= 19:
                    point = 18
                point -= 1
                oki_ishi = point // 2 + 1
                komi = 0
                if (point % 2) == 1:
                    komi = -6
                msg = '向 '
                if oki_ishi > 1:
                    msg = msg + zenhan.h2z(str.format('{:.1g}', oki_ishi)) + '子'
                if komi == 0:
                    msg = msg + 'コミなし'
                else:
                    msg = msg + '逆コミ６目'
            else:
                point = point * -1
                if point >= 19:
                    point = 18
                point -= 1
                oki_ishi = point // 2 + 1
                komi = 0
                if (point % 2) == 1:
                    komi = -6
                msg = ''
                if oki_ishi > 1:
                    msg = msg + zenhan.h2z(str.format('{:.1g}', oki_ishi)) + '子'
                if komi == 0:
                    msg = msg + 'コミなし'
                else:
                    msg = msg + '逆コミ６目'

        return '[' + str(handy_diff) + '] ' + msg

    # 対局者決定サブ
    def player_decision_sub(self, taisensha_info_rec):

        # 既に組み合わせされていたらスキップ
        if len(taisensha_info_rec.taisen_rireki) >= self.taisenNo:
            return

        # ペアを決定
        pair_rec = self.get_pair_kettei(taisensha_info_rec)

        if pair_rec == None:
            # ペアが見つからなかったら不戦勝扱いにする
            # 対戦リストに登録
            taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = taisensha_info_rec.name, pair_name = None, taisensha_name1 = '不戦勝', taisensha_name2 = None, kekka = '〇')
            taisensha_info_rec.taisen_rireki.append(taisen_rireki_info)
            return

        # 対戦者を決定
        aite_info = self.get_taisen_kettei(taisensha_info_rec, pair_rec)

        # 対戦者なし
        if aite_info == None:
            # 対戦相手が見つからなかったら不戦勝扱いにする
            # 対戦リストに登録
            taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = taisensha_info_rec.name, pair_name = None, taisensha_name1 = '不戦勝', taisensha_name2 = None, kekka = '〇')
            taisensha_info_rec.taisen_rireki.append(taisen_rireki_info)
            return

        # 対戦相手のペアを決定
        aite_pair_rec = self.get_aite_pair_kettei(taisensha_info_rec, pair_rec, aite_info)

        if aite_pair_rec == None:
            # 対戦相手のペアが見つからなかったら不戦勝扱いにする
            # 対戦リストに登録
            taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = taisensha_info_rec.name, pair_name = None, taisensha_name1 = '不戦勝', taisensha_name2 = None, kekka = '〇')
            taisensha_info_rec.taisen_rireki.append(taisen_rireki_info)
            return

        # 対戦リストに登録
        taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = taisensha_info_rec.name, pair_name = pair_rec.name, taisensha_name1 = aite_info.name, taisensha_name2 = aite_pair_rec.name, kekka = None)
        taisensha_info_rec.taisen_rireki.append(taisen_rireki_info)

        # ペアの対局リストに登録
        taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = pair_rec.name, pair_name = taisensha_info_rec.name, taisensha_name1 = aite_info.name, taisensha_name2 = aite_pair_rec.name, kekka = None)
        pair_rec.taisen_rireki.append(taisen_rireki_info)

        # 対戦相手１の対局リストに登録
        taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = aite_info.name, pair_name = aite_pair_rec.name, taisensha_name1 = taisensha_info_rec.name, taisensha_name2 = pair_rec.name, kekka = None)
        aite_info.taisen_rireki.append(taisen_rireki_info)

        # 対戦相手２の対局リストに登録
        taisen_rireki_info = taisen_rireki(no = self.taisenNo, name = aite_pair_rec.name, pair_name = aite_info.name, taisensha_name1 = taisensha_info_rec.name, taisensha_name2 = pair_rec.name, kekka = None)
        aite_pair_rec.taisen_rireki.append(taisen_rireki_info)

        return

    # 対局者決定
    def player_decision(self, taisenNo, execel_file_name, save_execel_file_name):

        self.taisenNo = taisenNo

        wb = openpyxl.load_workbook(execel_file_name)
        sheet = wb.active

        self.taisensha_info_list = []

        # エクセルから参加者と過去の対戦情報を読み取り
        self.read_excel(sheet)

        # 過去の対戦情報の矛盾をチェック
        if self.check_taisen_rireki(self.taisenNo - 1) == False:
            return

        # 不戦勝数、勝ち数、棋力＋ランダム数、登録順にマッチングする
        self.taisensha_info_list = sorted(self.taisensha_info_list, key=lambda x: (x.no_battles_cnt, x.score, x.kiryoku + x.kiryoku_random_val, x.random_row), reverse=True)

        # 相手ペアは不戦勝数、勝ち数(降順)、棋力(降順)＋ランダム数、登録順で見つける
        self.aite_pair_info_list = sorted(self.taisensha_info_list, key=lambda x: (x.score * -1, (x.kiryoku + x.kiryoku_random_val) * -1, x.random_row), reverse=True)

        # 対戦の組み合わせを計算
        for taisensha_info_rec in self.taisensha_info_list:

            self.player_decision_sub(taisensha_info_rec)

        # エクセルにペアと対戦者名とハンディを書き込み
        for rec in self.taisensha_info_list:
            if len(rec.taisen_rireki) >= self.taisenNo:
                senreki = rec.taisen_rireki[self.taisenNo - 1]
                col = self.taisen_start_col + (self.taisenNo - 1) * 4
                # ペア名
                sheet.cell(rec.row, col).value = senreki.pair_name
                # 対戦相手
                if senreki.taisensha_name2 != None:
                    sheet.cell(rec.row, col + 1).value = senreki.taisensha_name1 + '、' + senreki.taisensha_name2
                else:
                    sheet.cell(rec.row, col + 1).value = senreki.taisensha_name1
                # ハンディ
                if senreki.taisensha_name2 != None:
                    sheet.cell(rec.row, col + 2).value = self.get_handycap(senreki.name, senreki.pair_name, senreki.taisensha_name1, senreki.taisensha_name2)
                # 成績(不戦勝の場合)
                if senreki.kekka != None:
                    sheet.cell(rec.row, col + 3).value = senreki.kekka

        wb.save(save_execel_file_name)

    # 成績決定
    def write_result(self, execel_file_name, save_execel_file_name):

        wb = openpyxl.load_workbook(execel_file_name)
        sheet = wb.active

        # 成績の列位置を取得
        self.result_col_info(sheet)

        self.taisenNo = int((self.WIN_col - self.taisen_start_col) / 4)

        self.taisensha_info_list = []

        # エクセルから参加者と過去の対戦情報を読み取り
        self.read_excel(sheet)

        # 過去の対戦情報の矛盾をチェック
        if self.check_taisen_rireki(self.taisenNo) == False:
            return

        # SOSを計算
        for rec in self.taisensha_info_list:
            # SOS
            rec.sos = self.get_sos(rec)
            # SOSOS
            rec.sosos = self.get_sosos(rec)

        # スコア・SOS・SOSOS順にする
        self.taisensha_info_list = sorted(self.taisensha_info_list, key=lambda x: (x.score, x.sos, x.sosos), reverse=True)

        # 順位を計算
        current_jyuni = 1
        jyuni_count = 1
        old_rec = None
        for rec in self.taisensha_info_list:
            if old_rec != None:
                if old_rec.score != rec.score or old_rec.sos != rec.sos or old_rec.sosos != rec.sosos:
                    if self.pair_kotei_flag:
                        current_jyuni += 1
                    else:
                        current_jyuni += jyuni_count
                    jyuni_count = 1
                else:
                    jyuni_count += 1
            rec.jyuni = current_jyuni
            old_rec = rec

        # 結果を書き込み
        for rec in self.taisensha_info_list:
            # 勝ち数
            sheet.cell(rec.row, self.WIN_col).value = rec.score
            # SOS
            sheet.cell(rec.row, self.SOS_col).value = rec.sos
            # SOSOS
            sheet.cell(rec.row, self.SOSOS_col).value = rec.sosos
            # 順位
            sheet.cell(rec.row, self.JYUNI_col).value = rec.jyuni

        wb.save(save_execel_file_name)

if __name__ == "__main__":

    excel_file_name = 'ペア碁大会_対局者一覧_Test2.xlsx'
    save_execel_file_name = 'ペア碁大会_対局者一覧_結果.xlsx'
    # 'result'もしくは対戦番号(1～)
    cmd = '1'

    proc = PairMatchTable(seibetsu_flag = True, pair_kotei_flag = False, kiryoku_random_val_size = 3)

    if cmd == 'result':
        proc.write_result(excel_file_name, save_execel_file_name)
    else:
        taisenNo = int(cmd)
        proc.player_decision(taisenNo, excel_file_name, save_execel_file_name)


