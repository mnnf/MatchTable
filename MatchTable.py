import sys
import openpyxl
import datetime
import random
from collections import namedtuple
import zenhan

# 参加者のデータ構造体定義
class taisensha_info:
    def __init__(self, row, name, kiryoku, score, fusensho_count, random_seq, taisen_rireki, sos, sosos, jyuni):
        self.row = row
        self.name = name
        self.kiryoku = kiryoku
        self.score = score
        self.fusensho_count = fusensho_count
        self.random_seq = random_seq
        self.taisen_rireki = taisen_rireki
        self.sos = sos
        self.sosos = sosos
        self.jyuni = jyuni

# 参加者の対戦履歴のデータ構造体定義
class taisen_rireki:
    def __init__(self, no, name1, name2, kekka):
        self.no = no
        self.name1 = name1
        self.name2 = name2
        self.kekka = kekka

# 対戦相手のデータ構造体定義
class taisen_aite_info:
    def __init__(self, taisensha_info, taikyoku_su, index):
        self.taisensha_info = taisensha_info
        self.taikyoku_su = taikyoku_su
        self.index = index

# マッチングクラス
class MatchTable:

    def __init__(self):
        # タイトルの行位置
        self.title_row = 1

        # 参加者リストの開始行位置
        self.start_sankasha_row = 2

        # 参加者リストの名前列位置
        self.sankasha_name_col = 1

        # 対戦回の開始列位置
        self.taisen_start_col = 3

        # 勝ち数の列位置
        self.WIN_col = 0

        # SOSの列位置
        self.SOS_col = 0

        # SOSOSの列位置
        self.SOSOS_col = 0

        # 順位の列位置
        self.JYUNI_col = 0

        # 参加者リスト
        self.taisensha_info_list = []

        # 対戦番号(１～）
        self.taisenNo = 0

    # 棋力を取得(5D=5,1D=1,1K=0,2K=-1)
    def get_kiryoku(self, dat):
        kiryoku = 0
        if dat[1:2] == 'D':
            kiryoku = int(dat.replace('D', ''))
        else:
            kiryoku = 1 - int(dat.replace('K', ''))
        return kiryoku

    # 対戦未決定リスト取得
    def get_taisen_mikettei_list(self, taisensha_info_rec):

        mikettei_list = []

        for rec in self.taisensha_info_list:
            if rec.name == taisensha_info_rec.name:
                continue
            if len(rec.taisen_rireki) < self.taisenNo:
                mikettei_list.append(rec)

        return mikettei_list

    # 対戦未決定リスト取得
    def get_taisen_mikettei_list2(self, taisensha_info_rec):

        # 対戦者未決定リスト取得
        mikettei_list = self.get_taisen_mikettei_list(taisensha_info_rec)

        # 対戦者なし
        if len(mikettei_list) == 0:
            return mikettei_list

        mikettei_list2 = []

        for index, aite_info in enumerate(mikettei_list):
            # 過去対戦数を取得
            taikyoku_su = self.get_taikyoku_su(taisensha_info_rec.taisen_rireki, aite_info.name)
            mikettei_list2.append(taisen_aite_info(aite_info, taikyoku_su, index))

        # 対戦回数・登録順にする
        mikettei_list2 = sorted(mikettei_list2, key=lambda x: (x.taikyoku_su, x.index))

        mikettei_list3 = []

        for rec in mikettei_list2:
            mikettei_list3.append(rec.taisensha_info)

        return mikettei_list3

    # スコアを取得
    def get_score(self, taisen_rireki_list):

        score = 0
        for rec in taisen_rireki_list:
            if (rec.kekka == '〇'):
                score += 1
        return score

    # 不戦勝数を取得
    def get_fusensho_count(self, taisen_rireki_list):

        score = 0
        for rec in taisen_rireki_list:
            if rec.kekka == '〇' and rec.name2 == '不戦勝':
                score += 1
        return score

    # 過去の対局数
    def get_taikyoku_su(self, taisen_rireki_list, taisen_aite_name):
        taikyoku_su = 0
        for rec in taisen_rireki_list:
            if rec.name2 == taisen_aite_name:
                taikyoku_su += 1
        return taikyoku_su

    # 対戦相手の対戦者情報を取得
    def get_aite_info(self, name):
        for rec in self.taisensha_info_list:
            if rec.name == name:
                return rec
        return None

    # SOSを計算
    def get_sos(self, taisensha_info_rec):
        sos = 0
        for rec in taisensha_info_rec.taisen_rireki:
            aite_info = self.get_aite_info(rec.name2)
            if aite_info != None:
                score = self.get_score(aite_info.taisen_rireki)
                sos += score
        return sos

    # SOSOSを計算
    def get_sosos(self, taisensha_info_rec):
        sosos = 0
        for rec in taisensha_info_rec.taisen_rireki:
            aite_info = self.get_aite_info(rec.name2)
            if aite_info != None:
                sos = self.get_sos(aite_info)
                sosos += sos
        return sosos

    # 成績の列位置を取得
    def result_col_info(self, sheet):
        global WIN_col
        global SOS_col
        global SOSOS_col
        global JYUNI_col
        for col in range(1, sheet.max_column + 1):
            title = sheet.cell(self.title_row, col).value
            if title == '勝ち数':
                WIN_col = col
            if title == 'SOS':
                SOS_col = col
            if title == 'SOSOS':
                SOSOS_col = col
            if title == '順位':
                JYUNI_col = col

    # エクセルから参加者と過去の対戦情報を読み取り
    def read_excel(self, sheet):

        for row in range(self.start_sankasha_row, sheet.max_row + 1):

            # 対局者名を取得
            name = sheet.cell(row, self.sankasha_name_col).value
            if name != None:

                # 棋力を取得
                kiryoku = self.get_kiryoku(sheet.cell(row, self.sankasha_name_col + 1).value)

                # 対戦履歴情報取得
                taisen_rireki_info_list = []

                for taisenNo_loop in range(1, self.taisenNo + 1):
                    # 対戦者の名前と結果を取得
                    col = self.taisen_start_col + (taisenNo_loop - 1) * 3
                    taisensha_name = sheet.cell(row, col).value
                    kekka = sheet.cell(row, col + 2).value
                    if taisensha_name != None:
                        taisen_rireki_info = taisen_rireki(no = taisenNo_loop, name1 = name, name2 = taisensha_name, kekka = kekka)
                        taisen_rireki_info_list.append(taisen_rireki_info)

                # 参加者の勝ち星を取得
                score = self.get_score(taisen_rireki_info_list)

                # 参会者の不戦勝を取得
                fusensho_count = self.get_fusensho_count(taisen_rireki_info_list)

                # ランダム順位
                random_seq = random.randint(1, 100)

                # 参加者リストに追加
                self.taisensha_info_list.append(taisensha_info(row = row, name = name, kiryoku = kiryoku, score = score, fusensho_count = fusensho_count, random_seq = random_seq, taisen_rireki = taisen_rireki_info_list, sos = 0, sosos = 0, jyuni = 0))

    # 過去の対戦情報の矛盾をチェック
    def check_taisen_rireki(self, last_taisen_no):
        error_flag = False
        for rec in self.taisensha_info_list:
            for i, senreki in enumerate(rec.taisen_rireki):
                if i < last_taisen_no:
                    aite_info = self.get_aite_info(senreki.name2)
                    if aite_info == None:
                        # 不戦勝などは対局者情報が取得できない。
                        continue
                    if len(aite_info.taisen_rireki) > i:
                        aite_senreki = aite_info.taisen_rireki[i]
                        if senreki.kekka == aite_senreki.kekka:
                            print('{}回戦の {} vs {} 戦の結果が両者同じです。'.format(i+1, rec.name, aite_info.name))
                            error_flag = True
        return not error_flag

    # ハンディキャップ取得
    def get_handycap(self, name1, name2):

        kiryoku1 = self.get_aite_info(name1).kiryoku

        kiryoku2 = self.get_aite_info(name2).kiryoku

        handy_diff = kiryoku1 - kiryoku2

        msg = ''
        if handy_diff == 0:
            msg = '互先'
        else:
            point = (handy_diff / 0.5)
            if point > 0:
                if point >= 19:
                    point = 18
                point -= 1
                oki_ishi = point // 2 + 1
                komi = 0
                if (point % 2) == 1:
                    komi = -6
                msg = '向 '
                if oki_ishi > 1:
                    msg = msg + zenhan.h2z(str.format('{:.1g}', oki_ishi)) + '子'
                else:
                    msg = msg + '先'
            else:
                point = point * -1
                if point >= 19:
                    point = 18
                point -= 1
                oki_ishi = point // 2 + 1
                komi = 0
                if (point % 2) == 1:
                    komi = -6
                msg = ''
                if oki_ishi > 1:
                    msg = msg + zenhan.h2z(str.format('{:.1g}', oki_ishi)) + '子'
                else:
                    msg = msg + '先'

        return msg

    # 対局者決定
    def player_decision(self, taisenNo, execel_file_name, save_execel_file_name):

        self.taisenNo = taisenNo

        wb = openpyxl.load_workbook(execel_file_name)
        sheet = wb.active

        self.taisensha_info_list = []

        # エクセルから参加者と過去の対戦情報を読み取り
        self.read_excel(sheet)

        # 過去の対戦情報の矛盾をチェック
        if self.check_taisen_rireki(self.taisenNo - 1) == False:
            return

        # スコア・不戦勝数・棋力・ランダム順位・登録順にする
        self.taisensha_info_list = sorted(self.taisensha_info_list, key=lambda x: (x.score, x.fusensho_count, x.kiryoku, x.random_seq, x.row * -1), reverse=True)

        # 対戦の組み合わせを計算
        for i, rec in enumerate(self.taisensha_info_list):

            # 既に組み合わせされていたらスキップ
            if len(rec.taisen_rireki) >= self.taisenNo:
                continue

            # 対戦者未決定リスト取得
            mikettei_list = self.get_taisen_mikettei_list2(rec)

            # 対戦者なし
            if len(mikettei_list) == 0:
                # 不戦勝扱いにする
                # 対戦リストに登録
                taisen_rireki_info = taisen_rireki(self.taisenNo, rec.name, '不戦勝', '〇')
                rec.taisen_rireki.append(taisen_rireki_info)
            else:

                # 対戦相手決定
                aite_info = mikettei_list[0]

                # 対戦リストに登録
                taisen_rireki_info = taisen_rireki(self.taisenNo, rec.name, aite_info.name, None)
                rec.taisen_rireki.append(taisen_rireki_info)

                taisen_rireki_info = taisen_rireki(self.taisenNo, aite_info.name, rec.name, None)
                aite_info.taisen_rireki.append(taisen_rireki_info)

        # エクセルに対戦者名とハンディを書き込み
        for rec in self.taisensha_info_list:
            if len(rec.taisen_rireki) >= self.taisenNo:
                senreki = rec.taisen_rireki[self.taisenNo - 1]
                col = self.taisen_start_col + (self.taisenNo - 1) * 3
                # 対戦相手
                sheet.cell(rec.row, col).value = senreki.name2
                # ハンディ
                sheet.cell(rec.row, col + 1).value = self.get_handycap(senreki.name1, senreki.name2)
                # 成績(不戦勝の場合)
                if senreki.kekka != None:
                    col = self.taisen_start_col + (self.taisenNo - 1) * 3 + 2
                    sheet.cell(rec.row, col).value = senreki.kekka

        wb.save(save_execel_file_name)

    # 成績決定
    def write_result(self, execel_file_name, save_execel_file_name):

        wb = openpyxl.load_workbook(execel_file_name)
        sheet = wb.active

        # 成績の列位置を取得
        self.result_col_info(sheet)

        self.taisenNo = int((WIN_col - self.taisen_start_col) / 3)

        self.taisensha_info_list = []

        # エクセルから参加者と過去の対戦情報を読み取り
        self.read_excel(sheet)

        # 過去の対戦情報の矛盾をチェック
        if self.check_taisen_rireki(self.taisenNo) == False:
            return

        # SOSを計算
        for rec in self.taisensha_info_list:
            # SOS
            rec.sos = self.get_sos(rec)
            # SOSOS
            rec.sosos = self.get_sosos(rec)

        # スコア・SOS・SOSOS順にする
        self.taisensha_info_list = sorted(self.taisensha_info_list, key=lambda x: (x.score, x.sos, x.sosos), reverse=True)

        # 順位を計算
        current_jyuni = 1
        jyuni_count = 1
        old_rec = None
        for rec in self.taisensha_info_list:
            if old_rec != None:
                if old_rec.score != rec.score or old_rec.sos != rec.sos or old_rec.sosos != rec.sosos:
                    current_jyuni += jyuni_count
                    jyuni_count = 1
                else:
                    jyuni_count += 1
            rec.jyuni = current_jyuni
            old_rec = rec

        # 結果を書き込み
        for rec in self.taisensha_info_list:
            # 勝ち数
            sheet.cell(rec.row, WIN_col).value = rec.score
            # SOS
            sheet.cell(rec.row, SOS_col).value = rec.sos
            # SOSOS
            sheet.cell(rec.row, SOSOS_col).value = rec.sosos
            # 順位
            sheet.cell(rec.row, JYUNI_col).value = rec.jyuni

        wb.save(save_execel_file_name)

if __name__ == "__main__":

    excel_file_name = '対局者一覧サンプル_結果4.xlsx'
    save_execel_file_name = '対局者一覧サンプル_結果5.xlsx'
    # 'result'もしくは対戦番号(1～)
    cmd = 'result'

    proc = MatchTable()

    if cmd == 'result':
        proc.write_result(excel_file_name, save_execel_file_name)
    else:
        taisenNo = int(cmd)
        proc.player_decision(taisenNo, excel_file_name, save_execel_file_name)


